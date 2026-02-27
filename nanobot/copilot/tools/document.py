"""Document parsing tool: PDF, Excel, images, text files."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from nanobot.agent.tools.base import Tool

_MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB


class DocumentTool(Tool):
    """Tool for parsing documents: PDF, Excel, images, and text files."""

    def __init__(self, max_chars: int = 10000, max_file_size: int = _MAX_FILE_SIZE):
        self._max_chars = max_chars
        self._max_file_size = max_file_size

    @property
    def name(self) -> str:
        return "document"

    @property
    def description(self) -> str:
        return (
            "Parse and extract content from documents. "
            "Supports PDF (.pdf), Excel (.xlsx/.xls), images (.png/.jpg), "
            "and text files (.txt/.md/.csv/.json). "
            "Auto-detects format by extension."
        )

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "path": {
                    "type": "string",
                    "description": "Path to the document file",
                },
                "pages": {
                    "type": "string",
                    "description": "Page range for PDF (e.g. '1-5', '3'). Optional.",
                },
                "max_chars": {
                    "type": "integer",
                    "description": "Maximum characters to return (default 10000)",
                },
            },
            "required": ["path"],
        }

    async def execute(self, **kwargs: Any) -> str:
        file_path = kwargs.get("path", "")
        if not file_path:
            return "Error: file path required"

        path = Path(file_path)
        if not path.exists():
            return f"Error: file not found: {file_path}"

        size = path.stat().st_size
        if size > self._max_file_size:
            return f"Error: file is {size:,} bytes, exceeds {self._max_file_size:,} byte limit"

        max_chars = kwargs.get("max_chars", self._max_chars)
        ext = path.suffix.lower()

        try:
            if ext == ".pdf":
                return self._parse_pdf(path, kwargs.get("pages"), max_chars)
            elif ext in (".xlsx", ".xls"):
                return self._parse_excel(path, max_chars)
            elif ext in (".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"):
                return self._parse_image(path)
            elif ext in (".txt", ".md", ".csv", ".json", ".yaml", ".yml", ".log", ".xml", ".html"):
                return self._parse_text(path, max_chars)
            else:
                # Try as text
                return self._parse_text(path, max_chars)
        except Exception as e:
            return f"Error parsing {path.name}: {e}"

    @staticmethod
    def _parse_pdf(path: Path, pages: str | None, max_chars: int) -> str:
        """Parse PDF using PyMuPDF."""
        import fitz

        doc = fitz.open(str(path))
        total_pages = len(doc)

        # Determine page range
        start, end = 0, total_pages
        if pages:
            parts = pages.split("-")
            start = max(0, int(parts[0]) - 1)
            end = int(parts[-1]) if len(parts) > 1 else start + 1
            end = min(end, total_pages)

        text_parts = []
        for i in range(start, end):
            page = doc[i]
            text_parts.append(f"--- Page {i + 1} ---\n{page.get_text()}")

        doc.close()

        result = "\n".join(text_parts)
        metadata = f"PDF: {path.name} ({total_pages} pages, showing {start + 1}-{end})\n\n"

        if len(result) > max_chars:
            result = result[:max_chars] + "\n... (truncated)"

        return metadata + result

    @staticmethod
    def _parse_excel(path: Path, max_chars: int) -> str:
        """Parse Excel using openpyxl."""
        import openpyxl

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        parts = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            parts.append(f"--- Sheet: {sheet_name} ({len(rows)} rows) ---")

            # Header detection: use first row as headers
            headers = [str(c) if c is not None else "" for c in rows[0]]
            parts.append(" | ".join(headers))
            parts.append("-" * min(80, len(" | ".join(headers))))

            # Data rows (limit to 200)
            for row in rows[1:201]:
                cells = [str(c) if c is not None else "" for c in row]
                parts.append(" | ".join(cells))

            if len(rows) > 201:
                parts.append(f"... ({len(rows) - 201} more rows)")

        wb.close()

        result = "\n".join(parts)
        if len(result) > max_chars:
            result = result[:max_chars] + "\n... (truncated)"

        return f"Excel: {path.name}\n\n{result}"

    @staticmethod
    def _parse_image(path: Path) -> str:
        """Return image metadata (actual vision analysis is via LLM natively)."""
        import os
        size = os.path.getsize(path)
        return (
            f"Image: {path.name}\n"
            f"Size: {size:,} bytes\n"
            f"Format: {path.suffix.upper()}\n"
            f"Note: To analyze this image, include the file path in your message "
            f"and the LLM vision capability will process it directly."
        )

    @staticmethod
    def _parse_text(path: Path, max_chars: int) -> str:
        """Read text files directly."""
        content = path.read_text(encoding="utf-8", errors="replace")
        if len(content) > max_chars:
            content = content[:max_chars] + "\n... (truncated)"
        return f"File: {path.name} ({len(content)} chars)\n\n{content}"
