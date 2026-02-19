"""Context builder for assembling agent prompts."""

import base64
import mimetypes
import platform
from pathlib import Path
from typing import Any

from nanobot.agent.memory import MemoryStore
from nanobot.agent.skills import SkillsLoader


class ContextBuilder:
    """
    Builds the context (system prompt + messages) for the agent.
    
    Assembles bootstrap files, memory, skills, and conversation history
    into a coherent prompt for the LLM.
    """
    
    BOOTSTRAP_FILES = ["SOUL.md", "USER.md", "AGENTS.md", "POLICY.md"]
    
    def __init__(self, workspace: Path):
        self.workspace = workspace
        self.memory = MemoryStore(workspace)
        self.skills = SkillsLoader(workspace)
    
    def build_system_prompt(self, skill_names: list[str] | None = None) -> str:
        """
        Build the system prompt from bootstrap files, memory, and skills.
        
        Args:
            skill_names: Optional list of skills to include.
        
        Returns:
            Complete system prompt.
        """
        parts = []
        
        # Core identity
        parts.append(self._get_identity())
        
        # Bootstrap files
        bootstrap = self._load_bootstrap_files()
        if bootstrap:
            parts.append(bootstrap)
        
        # Memory context
        memory = self.memory.get_memory_context()
        if memory:
            parts.append(f"# Memory\n\n{memory}")
        
        # Skills - progressive loading
        # 1. Always-loaded skills: include full content
        always_skills = self.skills.get_always_skills()
        if always_skills:
            always_content = self.skills.load_skills_for_context(always_skills)
            if always_content:
                parts.append(f"# Active Skills\n\n{always_content}")
        
        # 2. Available skills: only show summary (agent uses read_file to load)
        skills_summary = self.skills.build_skills_summary()
        if skills_summary:
            parts.append(f"""# Skills

The following skills extend your capabilities. To use a skill, read its SKILL.md file using the read_file tool.
Skills with available="false" need dependencies installed first - you can try installing them with apt/brew.

{skills_summary}""")
        
        return "\n\n---\n\n".join(parts)
    
    def _get_identity(self) -> str:
        """Get the core identity section."""
        from datetime import datetime
        import time as _time
        now = datetime.now().strftime("%Y-%m-%d %H:%M (%A)")
        tz = _time.strftime("%Z") or "UTC"
        workspace_path = str(self.workspace.expanduser().resolve())
        system = platform.system()
        runtime = f"{'macOS' if system == 'Darwin' else system} {platform.machine()}, Python {platform.python_version()}"
        
        return f"""# System Context

See SOUL.md for your identity and behavioral principles.
See USER.md for user profile and preferences.
See AGENTS.md for operational rules and tools.

## Current Time
{now} ({tz})

## Runtime
{runtime}

## Workspace
Your workspace is at: {workspace_path}
- Identity files: SOUL.md, USER.md, AGENTS.md (you can edit these)
- Memory system manual: MEMORY.md
- Consolidation memory: {workspace_path}/memory/MEMORY.md
- History log: {workspace_path}/memory/HISTORY.md (grep-searchable)
- Custom skills: {workspace_path}/skills/{{skill-name}}/SKILL.md

IMPORTANT: When responding to direct questions or conversations, reply directly with your text response.
Only use the 'message' tool when you need to send a message to a specific chat channel (like WhatsApp).
For normal conversation, just respond with text - do not call the message tool."""
    
    def _load_bootstrap_files(self) -> str:
        """Load all bootstrap files from workspace."""
        parts = []
        
        for filename in self.BOOTSTRAP_FILES:
            file_path = self.workspace / filename
            if file_path.exists():
                content = file_path.read_text(encoding="utf-8")
                parts.append(f"## {filename}\n\n{content}")
        
        return "\n\n".join(parts) if parts else ""
    
    _ONBOARDING_PROMPT = """

## ONBOARDING INTERVIEW MODE

You are conducting a getting-to-know-you interview with your user. This is your chance to learn everything you need to be maximally helpful.

**Rules:**
- Ask ONE question at a time, wait for the answer
- Be warm and conversational, not robotic
- Follow up naturally if an answer is interesting or unclear
- Track which section you're on

**Sections to cover (in order):**
1. BASICS: Name, timezone, languages they speak
2. LIFE CONTEXT: Current life situation, typical day/week, main responsibilities
3. GOALS: Biggest goals (career/personal/financial/health), 6-month priority, things they keep putting off
4. PROJECTS: 2-3 active projects, which needs help, upcoming deadlines
5. WORK STYLE: Brief vs detailed responses, proactive vs reactive, active hours vs DND times
6. AUTONOMY: When to act vs ask — hypothetical scenarios like calendar conflicts, complex tasks, what to NEVER decide alone, what to handle freely
7. ASSISTANCE: Energy-draining tasks to hand off, recurring reminders to track, how to deliver bad news/problems, anything else

**When you finish ALL sections, write to these files:**
1. **{workspace}/USER.md** — LEAN profile (~10 lines): name, timezone, language, communication style, key autonomy rules. Loaded every prompt, keep it tight.
2. **{workspace}/SOUL.md** — Update your behavioral principles based on what you learned about how the user wants you to operate (proactive vs reactive, autonomy level, communication style).
3. **{workspace}/AGENTS.md** — Update operational rules: what tools/integrations are available, what's NOT configured, scheduling preferences, escalation rules.
4. **{workspace}/memory/MEMORY.md** — LEAN working snapshot (~300 words max): active goals, current projects, immediate priorities. NOT a detailed store — this is injected every prompt. Detail goes to HISTORY.md.
5. **{workspace}/memory/HISTORY.md** — Append a detailed summary of the onboarding conversation: full context, preferences discussed, action plans, everything that doesn't fit in the lean files above.
6. Summarize what you learned and present your action plan to the user.
7. Tell the user the interview is complete and they can start chatting normally.

Start by introducing yourself warmly and asking the first question.
"""

    def build_messages(
        self,
        history: list[dict[str, Any]],
        current_message: str,
        skill_names: list[str] | None = None,
        media: list[str] | None = None,
        channel: str | None = None,
        chat_id: str | None = None,
        session_metadata: dict[str, Any] | None = None,
    ) -> list[dict[str, Any]]:
        """
        Build the complete message list for an LLM call.

        Args:
            history: Previous conversation messages.
            current_message: The new user message.
            skill_names: Optional skills to include.
            media: Optional list of local file paths for images/media.
            channel: Current channel (telegram, feishu, etc.).
            chat_id: Current chat/user ID.
            session_metadata: Optional session metadata (for onboarding etc.).

        Returns:
            List of messages including system prompt.
        """
        messages = []

        # System prompt
        system_prompt = self.build_system_prompt(skill_names)
        if channel and chat_id:
            system_prompt += f"\n\n## Current Session\nChannel: {channel}\nChat ID: {chat_id}"

        # Inject onboarding interview prompt when active
        if session_metadata and session_metadata.get("onboarding_active"):
            workspace_path = str(self.workspace.expanduser().resolve())
            system_prompt += self._ONBOARDING_PROMPT.replace("{workspace}", workspace_path)

        messages.append({"role": "system", "content": system_prompt})

        # History
        messages.extend(history)

        # Current message (with optional image attachments)
        user_content = self._build_user_content(current_message, media)
        messages.append({"role": "user", "content": user_content})

        return messages

    def _build_user_content(self, text: str, media: list[str] | None) -> str | list[dict[str, Any]]:
        """Build user message content with optional media (images + documents)."""
        if not media:
            return text

        images = []
        for file_path in media:
            p = Path(file_path)
            if not p.is_file():
                continue

            mime, _ = mimetypes.guess_type(file_path)
            ext = p.suffix.lower()

            # Images: base64-encode for LLM vision
            if mime and mime.startswith("image/"):
                b64 = base64.b64encode(p.read_bytes()).decode()
                images.append({"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}})
                continue

            # Non-image files: extract text and append to message
            extracted = self._extract_document_text(p, ext)
            if extracted:
                text += f"\n\n{extracted}"

        if not images:
            return text
        return images + [{"type": "text", "text": text}]

    @staticmethod
    def _extract_document_text(p: Path, ext: str) -> str:
        """Extract text from a document file for inline context injection."""
        max_chars = 10000

        try:
            if ext == ".pdf":
                import fitz
                doc = fitz.open(str(p))
                parts = [f"--- Page {i+1} ---\n{doc[i].get_text()}" for i in range(len(doc))]
                doc.close()
                result = "\n".join(parts)
                if len(result) > max_chars:
                    result = result[:max_chars] + "\n... (truncated)"
                return f"[Document: {p.name}]\n{result}"

            if ext in (".xlsx", ".xls"):
                import openpyxl
                wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
                parts = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows:
                        continue
                    parts.append(f"--- Sheet: {sheet_name} ({len(rows)} rows) ---")
                    for row in rows[:201]:
                        parts.append(" | ".join(str(c) if c is not None else "" for c in row))
                    if len(rows) > 201:
                        parts.append(f"... ({len(rows) - 201} more rows)")
                wb.close()
                result = "\n".join(parts)
                if len(result) > max_chars:
                    result = result[:max_chars] + "\n... (truncated)"
                return f"[Document: {p.name}]\n{result}"

            # Text-like files
            if ext in (".txt", ".md", ".csv", ".json", ".xml", ".html", ".log", ".yaml", ".yml"):
                content = p.read_text(encoding="utf-8", errors="replace")
                if len(content) > max_chars:
                    content = content[:max_chars] + "\n... (truncated)"
                return f"[Document: {p.name}]\n{content}"

            # Unknown extension: try reading as text
            content = p.read_text(encoding="utf-8", errors="strict")
            if len(content) > max_chars:
                content = content[:max_chars] + "\n... (truncated)"
            return f"[Document: {p.name}]\n{content}"

        except ImportError:
            return f"[Document: {p.name} — required library not installed]"
        except UnicodeDecodeError:
            return f"[Unsupported file: {p.name}]"
        except Exception as e:
            return f"[Document: {p.name} — extraction failed: {e}]"
    
    def add_tool_result(
        self,
        messages: list[dict[str, Any]],
        tool_call_id: str,
        tool_name: str,
        result: str
    ) -> list[dict[str, Any]]:
        """
        Add a tool result to the message list.
        
        Args:
            messages: Current message list.
            tool_call_id: ID of the tool call.
            tool_name: Name of the tool.
            result: Tool execution result.
        
        Returns:
            Updated message list.
        """
        messages.append({
            "role": "tool",
            "tool_call_id": tool_call_id,
            "name": tool_name,
            "content": result
        })
        return messages
    
    def add_assistant_message(
        self,
        messages: list[dict[str, Any]],
        content: str | None,
        tool_calls: list[dict[str, Any]] | None = None,
        reasoning_content: str | None = None,
    ) -> list[dict[str, Any]]:
        """
        Add an assistant message to the message list.
        
        Args:
            messages: Current message list.
            content: Message content.
            tool_calls: Optional tool calls.
            reasoning_content: Thinking output (Kimi, DeepSeek-R1, etc.).
        
        Returns:
            Updated message list.
        """
        msg: dict[str, Any] = {"role": "assistant", "content": content or ""}
        
        if tool_calls:
            msg["tool_calls"] = tool_calls
        
        # Thinking models reject history without this
        if reasoning_content:
            msg["reasoning_content"] = reasoning_content
        
        messages.append(msg)
        return messages
